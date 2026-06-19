"""Excel ingestion + validation.

Validation rules for an uploaded input file:
  - size <= 20 MB                          (enforced in the router, pre-read)
  - file opens as a real .xlsx             (corruption check)
  - the sheet has a header row + data      (no-data check)
  - no merged cells in/around the header   (catches "two headers merged together")
  - header cells are non-empty and unique  (clean, mappable headers)
"""

import zipfile
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MAX_BYTES = 20 * 1024 * 1024  # 20 MB compressed (enforced in the router)
# An .xlsx is a zip; a small file can inflate to gigabytes ("zip bomb"). Reject
# anything whose uncompressed payload, or expansion ratio, is implausible for a
# real spreadsheet before openpyxl ever parses it.
MAX_UNCOMPRESSED = 400 * 1024 * 1024  # 400 MB total inflated
MAX_COMPRESSION_RATIO = 200
# Structural caps so a (valid) but enormous sheet can't exhaust memory.
MAX_DATA_ROWS = 1_000_000
MAX_COLUMNS = 1_000


def _cell_to_str(v) -> str:
    """Convert any cell value to a JSON-safe string the cleaner can parse."""
    import datetime as _dt

    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        return v.isoformat(sep=" ")
    if isinstance(v, (_dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))  # avoid 8.9e+12 style for barcodes
    return str(v)


class ExcelValidationError(Exception):
    """Raised when an uploaded file fails a validation rule."""

    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(message)


@dataclass
class SheetInfo:
    sheet_name: str
    header_row: int
    headers: list[str]
    n_columns: int
    n_rows: int  # data rows, excluding the header
    rows: list[list] = field(default_factory=list)  # extracted data (strings)
    warnings: list[str] = field(default_factory=list)


def _first_data_row(ws) -> int:
    """Index of the first row that has any non-empty cell (the header row)."""
    for row in ws.iter_rows():
        if any(c.value not in (None, "") for c in row):
            return row[0].row
    return 0


def _guard_zip_bomb(source) -> None:
    """Reject decompression bombs before openpyxl parses the workbook."""
    try:
        with zipfile.ZipFile(source) as zf:
            total_uncompressed = 0
            total_compressed = 0
            for info in zf.infolist():
                total_uncompressed += info.file_size
                total_compressed += info.compress_size
            if total_uncompressed > MAX_UNCOMPRESSED:
                raise ExcelValidationError(
                    "too_large",
                    "The file expands to an unreasonable size when opened and was "
                    "rejected as a potential decompression bomb.",
                )
            if (
                total_compressed > 0
                and total_uncompressed / total_compressed > MAX_COMPRESSION_RATIO
            ):
                raise ExcelValidationError(
                    "too_large",
                    "The file has an abnormal compression ratio and was rejected "
                    "as a potential decompression bomb.",
                )
    except zipfile.BadZipFile:
        raise ExcelValidationError(
            "corrupted",
            "The file could not be opened. It may be corrupted or not a valid "
            ".xlsx file.",
        )
    finally:
        if hasattr(source, "seek"):
            source.seek(0)


def read_and_validate(source) -> SheetInfo:
    # --- decompression-bomb guard (before any XML parsing) -----------------
    _guard_zip_bomb(source)

    # --- corruption / unreadable -------------------------------------------
    try:
        wb = load_workbook(source, read_only=False, data_only=True)
    except Exception as exc:  # openpyxl raises various errors on bad files
        raise ExcelValidationError(
            "corrupted",
            "The file could not be opened. It may be corrupted or not a valid "
            f".xlsx file ({type(exc).__name__}).",
        )

    ws = wb.active
    if ws is None:
        raise ExcelValidationError("no_data", "The workbook has no sheets.")

    header_row = _first_data_row(ws)
    if header_row == 0:
        raise ExcelValidationError("no_data", "The sheet is completely empty.")

    # --- merged cells in the header band -----------------------------------
    # A merged range that touches the header row means two or more headers were
    # merged into one cell ("2 headers merged together").
    merged_in_header = [
        str(rng)
        for rng in ws.merged_cells.ranges
        if rng.min_row <= header_row <= rng.max_row
    ]
    if merged_in_header:
        raise ExcelValidationError(
            "merged_header",
            "Merged cells were found in the header row "
            f"({', '.join(merged_in_header)}). Please unmerge them so each "
            "column has its own header.",
        )

    # --- extract + validate header cells -----------------------------------
    headers: list[str] = []
    empty_positions: list[str] = []
    for cell in ws[header_row]:
        value = cell.value
        if value is None or str(value).strip() == "":
            # Trailing empty cells are fine; gaps between headers are not.
            empty_positions.append(get_column_letter(cell.column))
            headers.append("")
        else:
            headers.append(str(value).strip())

    # Drop trailing blanks, then ensure no internal gaps remain.
    while headers and headers[-1] == "":
        headers.pop()
    if "" in headers:
        gap_cols = [get_column_letter(i + 1) for i, h in enumerate(headers) if h == ""]
        raise ExcelValidationError(
            "empty_header",
            f"Some columns have empty headers (columns {', '.join(gap_cols)}). "
            "Every column with data needs a header.",
        )
    if not headers:
        raise ExcelValidationError("no_data", "No column headers were found.")
    if len(headers) > MAX_COLUMNS:
        raise ExcelValidationError(
            "too_large",
            f"The sheet has more than {MAX_COLUMNS} columns, which is not supported.",
        )

    # --- extract data rows (JSON-safe strings) ------------------------------
    n_cols = len(headers)
    data_rows: list[list] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue  # skip fully blank rows
        cells = list(row[:n_cols]) + [None] * (n_cols - len(row[:n_cols]))
        data_rows.append([_cell_to_str(v) for v in cells])
        if len(data_rows) > MAX_DATA_ROWS:
            raise ExcelValidationError(
                "too_large",
                f"The sheet has more than {MAX_DATA_ROWS:,} data rows, which is "
                "not supported.",
            )
    if not data_rows:
        raise ExcelValidationError(
            "no_data", "The file has headers but no data rows."
        )

    # --- duplicate headers (warning, not fatal) ----------------------------
    warnings: list[str] = []
    seen: dict[str, int] = {}
    for h in headers:
        seen[h] = seen.get(h, 0) + 1
    dupes = [h for h, n in seen.items() if n > 1]
    if dupes:
        warnings.append(f"Duplicate headers found: {', '.join(dupes)}.")

    return SheetInfo(
        sheet_name=ws.title,
        header_row=header_row,
        headers=headers,
        n_columns=len(headers),
        n_rows=len(data_rows),
        rows=data_rows,
        warnings=warnings,
    )
