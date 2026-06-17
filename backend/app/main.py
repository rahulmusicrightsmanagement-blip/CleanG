import os
from contextlib import asynccontextmanager

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from sqlalchemy import select, text

from .config import get_settings
from .database import Base, SessionLocal, engine
from .models import MasterColumn, User, UserRole
from .routers import auth, branches, clean, files, master, users
from .security import hash_password

settings = get_settings()

MASTER_FILE = os.path.join(os.path.dirname(__file__), "data", "master_output_format.xlsx")


def _seed_master_columns(db) -> None:
    """Load the canonical output schema from the bundled master workbook."""
    if db.scalar(select(MasterColumn).limit(1)) is not None:
        return
    wb = load_workbook(MASTER_FILE, data_only=True)
    ws = wb.active
    position = 0
    for cell in ws[1]:
        value = cell.value
        if value is None or str(value).strip() == "":
            continue
        position += 1
        db.add(MasterColumn(position=position, name=str(value).strip()))
    db.commit()


def _migrate(db) -> None:
    """Add the review-overlay columns to pre-existing tables (idempotent).

    `create_all` only creates missing tables, never alters existing ones, so the
    `corrections`/`dropped` columns are added here for databases created before
    cleaning moved fully in-memory.
    """
    db.execute(text(
        "ALTER TABLE uploaded_files "
        "ADD COLUMN IF NOT EXISTS corrections JSONB NOT NULL DEFAULT '{}'::jsonb"
    ))
    db.execute(text(
        "ALTER TABLE uploaded_files "
        "ADD COLUMN IF NOT EXISTS dropped JSONB NOT NULL DEFAULT '[]'::jsonb"
    ))
    # Cleaned rows are no longer persisted — drop the legacy table so its stale
    # rows can't block file/branch deletion via the old foreign key.
    db.execute(text("DROP TABLE IF EXISTS cleaned_rows"))
    db.commit()


def init_db() -> None:
    """Create tables and seed the bootstrap admin + master schema if missing."""
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        _migrate(db)
        has_user = db.scalar(select(User).limit(1))
        if has_user is None:
            db.add(
                User(
                    email=settings.admin_email,
                    full_name=settings.admin_name,
                    hashed_password=hash_password(settings.admin_password),
                    role=UserRole.admin,
                )
            )
            db.commit()
        _seed_master_columns(db)
    finally:
        db.close()


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(title="MRM Cleanser API", version="0.1.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origin_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(auth.router)
app.include_router(users.router)
app.include_router(branches.router)
app.include_router(master.router)
app.include_router(files.router)
app.include_router(clean.router)


@app.get("/api/health")
def health():
    return {"status": "ok"}
